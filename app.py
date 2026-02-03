import os
import sys
import tempfile
import threading
import uuid

from flask import Flask, render_template, request, send_file, flash, redirect, url_for, jsonify
import openpyxl
from werkzeug.utils import secure_filename


def get_resource_path(relative_path):
    """PyInstallerでパッケージ化された場合のリソースパスを取得"""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), relative_path)


app = Flask(__name__, template_folder=get_resource_path('templates'))
app.secret_key = os.urandom(24)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB制限（デスクトップアプリ向けに拡大）

ALLOWED_EXTENSIONS = {'xlsx', 'xlsm'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def copy_sheets(source_wb, sheet_names):
    """指定されたシートを新しいワークブックにコピー"""
    new_wb = openpyxl.Workbook()
    new_wb.remove(new_wb.active)

    for sheet_name in sheet_names:
        source_sheet = source_wb[sheet_name]
        new_sheet = new_wb.create_sheet(title=sheet_name)

        for row in source_sheet.iter_rows():
            for cell in row:
                new_cell = new_sheet[cell.coordinate]
                new_cell.value = cell.value

                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.number_format = cell.number_format
                    new_cell.alignment = cell.alignment.copy()

        for col in source_sheet.column_dimensions:
            new_sheet.column_dimensions[col].width = \
                source_sheet.column_dimensions[col].width

        for merged_cell in source_sheet.merged_cells.ranges:
            new_sheet.merge_cells(str(merged_cell))

    return new_wb


def copy_sheets_by_keyword(source_file, output_file, keyword):
    """キーワードを含むシートのみを新しいExcelファイルにコピー"""
    source_wb = openpyxl.load_workbook(source_file)
    matching_sheets = [name for name in source_wb.sheetnames if keyword in name]

    if not matching_sheets:
        return None, 0

    new_wb = copy_sheets(source_wb, matching_sheets)
    new_wb.save(output_file)
    return output_file, len(matching_sheets)


def copy_sheets_by_selection(source_file, output_file, selected_sheets):
    """選択されたシートを新しいExcelファイルにコピー"""
    source_wb = openpyxl.load_workbook(source_file)
    valid_sheets = [name for name in selected_sheets if name in source_wb.sheetnames]

    if not valid_sheets:
        return None, 0

    new_wb = copy_sheets(source_wb, valid_sheets)
    new_wb.save(output_file)
    return output_file, len(valid_sheets)


@app.route('/get_sheets', methods=['POST'])
def get_sheets():
    """アップロードされたExcelファイルのシート名一覧を返す"""
    if 'file' not in request.files:
        return jsonify({'error': 'ファイルが選択されていません'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'ファイルが選択されていません'}), 400

    if not allowed_file(file.filename):
        return jsonify({'error': '対応しているファイル形式は .xlsx, .xlsm です'}), 400

    try:
        wb = openpyxl.load_workbook(file, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        return jsonify({'sheets': sheets, 'count': len(sheets)})
    except Exception as e:
        return jsonify({'error': f'ファイルの読み込みに失敗しました: {str(e)}'}), 400


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('ファイルが選択されていません', 'error')
            return redirect(request.url)

        file = request.files['file']
        keyword = request.form.get('keyword', '').strip()
        selected_sheets = request.form.getlist('selected_sheets')
        extract_mode = request.form.get('extract_mode', 'keyword')

        if file.filename == '':
            flash('ファイルが選択されていません', 'error')
            return redirect(request.url)

        if not file or not allowed_file(file.filename):
            flash('対応しているファイル形式は .xlsx, .xlsm です', 'error')
            return redirect(request.url)

        if extract_mode == 'keyword' and not keyword:
            flash('検索キーワードを入力してください', 'error')
            return redirect(request.url)
        elif extract_mode == 'selection' and not selected_sheets:
            flash('抽出するシートを選択してください', 'error')
            return redirect(request.url)

        temp_dir = tempfile.mkdtemp()
        input_path = os.path.join(temp_dir, secure_filename(file.filename))
        file.save(input_path)

        try:
            if extract_mode == 'selection':
                output_filename = "抽出_選択シート.xlsx"
                output_path = os.path.join(temp_dir, output_filename)
                result_path, sheet_count = copy_sheets_by_selection(input_path, output_path, selected_sheets)
                error_msg = "選択されたシートが見つかりませんでした"
            else:
                output_filename = f"抽出_{keyword}.xlsx"
                output_path = os.path.join(temp_dir, output_filename)
                result_path, sheet_count = copy_sheets_by_keyword(input_path, output_path, keyword)
                error_msg = f"'{keyword}'を含むシートが見つかりませんでした"

            if result_path is None:
                flash(error_msg, 'warning')
                return redirect(request.url)

            return send_file(
                output_path,
                as_attachment=True,
                download_name=output_filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as e:
            flash(f'処理中にエラーが発生しました: {str(e)}', 'error')
            return redirect(request.url)

    return render_template('index.html')


def start_flask_server():
    """Flaskサーバーをバックグラウンドで起動"""
    app.run(host='127.0.0.1', port=5000, debug=False, threaded=True, use_reloader=False)


def main():
    """メインエントリーポイント: pywebviewでデスクトップアプリとして起動"""
    try:
        import webview

        # Flaskサーバーをバックグラウンドスレッドで起動
        server_thread = threading.Thread(target=start_flask_server, daemon=True)
        server_thread.start()

        # サーバーの起動を待機
        import time
        time.sleep(1)

        # pywebviewでウィンドウを作成
        webview.create_window(
            'SheetPick - Excelシート分離ツール',
            'http://127.0.0.1:5000',
            width=700,
            height=800,
            resizable=True,
            min_size=(500, 600)
        )
        webview.start()

    except ImportError:
        # pywebviewがインストールされていない場合はブラウザで開く
        import webbrowser

        print("pywebviewがインストールされていません。ブラウザで開きます。")
        print("アプリを終了するには Ctrl+C を押してください。")

        # サーバー起動前にブラウザを開く準備
        def open_browser():
            import time
            time.sleep(1.5)
            webbrowser.open('http://127.0.0.1:5000')

        browser_thread = threading.Thread(target=open_browser, daemon=True)
        browser_thread.start()

        # Flaskサーバーを起動（メインスレッド）
        app.run(host='127.0.0.1', port=5000, debug=False, threaded=True)


if __name__ == '__main__':
    main()
